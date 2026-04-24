import csv
import hashlib
import threading
from collections import Counter
from datetime import datetime, timedelta

from library_app.config import (
    DEFAULT_STUDENTS_FILE,
    DUPLICATE_SCAN_GAP_SECONDS,
    EXCEL_STUDENTS_FILE,
    LIBRARY_DB_FILE,
    LIBRARY_DATA_FILE,
    VISIT_FIELDS,
    VISITS_FILE,
)
from library_app.database import (
    create_visit,
    ensure_database_ready,
    fetch_open_visit,
    fetch_latest_visit_for_student,
    fetch_visits,
    update_visit_exit,
)
from library_app.time_utils import current_date_text, now_local, parse_local_timestamp, today_local

_STUDENT_CACHE = {}
_STUDENT_CACHE_SIGNATURE = ""
_STUDENT_CACHE_LOCK = threading.Lock()
_ANALYTICS_CACHE = {"signature": "", "token": "", "visits": []}
_ANALYTICS_CACHE_LOCK = threading.Lock()


def _file_signature(path):
    if not path.exists():
        return ""
    stat = path.stat()
    return f"{path.resolve()}::{stat.st_size}::{stat.st_mtime_ns}"


def _excel_source_files():
    base_dir = EXCEL_STUDENTS_FILE.parent
    preferred = []
    if EXCEL_STUDENTS_FILE.exists():
        preferred.append(EXCEL_STUDENTS_FILE)
    others = sorted([path for path in base_dir.glob("*.xlsx") if path.name != EXCEL_STUDENTS_FILE.name])
    return preferred + others


def _students_source_signature():
    excel_files = [path for path in _excel_source_files() if path.exists()]
    if excel_files:
        return "|".join(_file_signature(path) for path in excel_files)
    return _file_signature(get_students_file())


def _analytics_source_signature():
    return f"{_file_signature(LIBRARY_DB_FILE)}|{_students_source_signature()}"


def _normalize_student_row(row):
    return {
        "student_id": (row.get("student_id") or row.get("Student ID") or "").strip(),
        "name": (row.get("name") or row.get("Name") or "").strip(),
        "father_name": (row.get("father_name") or row.get("Father Name") or row.get("FATHER NAME") or "").strip(),
        "course": (
            row.get("course")
            or row.get("Course")
            or row.get("coursev1")
            or row.get("Coursev1")
            or row.get("branch")
            or row.get("Branch")
            or ""
        ).strip(),
        "phone": (row.get("phone") or row.get("Phone") or "").strip(),
        "valid_until": (row.get("valid_until") or row.get("Valid Until") or "").strip(),
    }


def _load_students_from_excel():
    from openpyxl import load_workbook

    students = {}
    for excel_file in _excel_source_files():
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = worksheet.iter_rows(values_only=True)
                headers = None
                for row in rows:
                    values = [str(cell).strip() if cell is not None else "" for cell in row]
                    if not any(values):
                        continue
                    normalized = [value.lower() for value in values]
                    if headers is None:
                        if (
                            ("name" in normalized or "name " in normalized or "studente name" in normalized or "students name" in normalized)
                            and "branch" in normalized
                            and "code" in normalized
                        ):
                            headers = values
                        continue

                    row_map = dict(zip(headers, values))
                    student_id = str(row_map.get("CODE", "")).strip()
                    name = str(
                        row_map.get("Name ", "")
                        or row_map.get("Name", "")
                        or row_map.get("STUDENTE NAME", "")
                        or row_map.get("Students Name", "")
                    ).strip()
                    if not student_id or not name:
                        continue
                    students[student_id] = {
                        "student_id": student_id,
                        "name": name,
                        "father_name": str(row_map.get("FATHER NAME", "") or row_map.get("Father Name", "")).strip(),
                        "course": str(row_map.get("BRANCH", "") or row_map.get("Branch", "")).strip(),
                        "phone": "",
                        "valid_until": "",
                    }
        finally:
            workbook.close()
    return students


def _load_students_from_csv():
    students = {}
    source = get_students_file()
    if source.suffix.lower() == ".xlsx" or not source.exists():
        return students
    with source.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            student = _normalize_student_row(row)
            if not student["student_id"]:
                continue
            students[student["student_id"]] = student
    return students


def _load_students_from_source():
    excel_files = [path for path in _excel_source_files() if path.exists()]
    if excel_files:
        return _load_students_from_excel()
    return _load_students_from_csv()


def get_students_file():
    if EXCEL_STUDENTS_FILE.exists():
        return EXCEL_STUDENTS_FILE
    if LIBRARY_DATA_FILE.exists():
        return LIBRARY_DATA_FILE
    return DEFAULT_STUDENTS_FILE


def ensure_students_file():
    ensure_database_ready()


def ensure_visits_file():
    ensure_database_ready()


def load_students():
    global _STUDENT_CACHE
    global _STUDENT_CACHE_SIGNATURE

    signature = _students_source_signature()
    if signature and signature == _STUDENT_CACHE_SIGNATURE and _STUDENT_CACHE:
        return _STUDENT_CACHE

    with _STUDENT_CACHE_LOCK:
        signature = _students_source_signature()
        if signature and signature == _STUDENT_CACHE_SIGNATURE and _STUDENT_CACHE:
            return _STUDENT_CACHE
        _STUDENT_CACHE = _load_students_from_source()
        _STUDENT_CACHE_SIGNATURE = signature
        return _STUDENT_CACHE


def load_visits():
    return fetch_visits()


def save_visits(visits):
    import csv

    with VISITS_FILE.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=VISIT_FIELDS)
        writer.writeheader()
        writer.writerows(visits)
    return visits


def is_membership_valid(student):
    valid_until = student.get("valid_until", "").strip()
    if not valid_until:
        return True, ""

    try:
        expiry_date = datetime.strptime(valid_until, "%Y-%m-%d").date()
    except ValueError:
        return False, "Student date format invalid"

    today = today_local()
    if today > expiry_date:
        return False, f"ID expired on {expiry_date.isoformat()}"

    return True, ""


def find_open_visit(visits, student_id, visit_date):
    for visit in reversed(visits):
        if (
            visit["student_id"] == student_id
            and visit["date"] == visit_date
            and not visit["exit_time"].strip()
        ):
            return visit
    return None


def parse_timestamp(date_text, time_text):
    return parse_local_timestamp(date_text, time_text)


def get_last_scan_timestamp(visits, student_id):
    for visit in reversed(visits):
        if visit["student_id"] != student_id:
            continue
        exit_timestamp = parse_timestamp(visit["date"], visit["exit_time"])
        if exit_timestamp is not None:
            return exit_timestamp
        entry_timestamp = parse_timestamp(visit["date"], visit["entry_time"])
        if entry_timestamp is not None:
            return entry_timestamp
    return None


def process_scan_result(student_id):
    student_id = str(student_id).strip()
    now = now_local()
    today = now.date().isoformat()
    student = load_students().get(student_id)

    if student is None:
        return {
            "ok": False,
            "message": f"Student ID not found: {student_id}",
            "student": None,
            "visit": None,
            "action": "not_found",
        }

    is_valid, reason = is_membership_valid(student)
    if not is_valid:
        return {
            "ok": False,
            "message": reason,
            "student": student,
            "visit": None,
            "action": "invalid",
        }

    last_visit = fetch_latest_visit_for_student(student_id)
    last_scan_timestamp = None
    if last_visit is not None:
        last_scan_timestamp = parse_timestamp(last_visit["date"], last_visit["exit_time"]) or parse_timestamp(
            last_visit["date"], last_visit["entry_time"]
        )
    if last_scan_timestamp is not None:
        if last_scan_timestamp.tzinfo is None:
            last_scan_timestamp = last_scan_timestamp.replace(tzinfo=now.tzinfo)
        elapsed = (now - last_scan_timestamp).total_seconds()
        # A future timestamp can happen if an older record was written with a mismatched
        # server clock or timezone. In that case, don't turn it into a nonsense cooldown.
        if elapsed < 0:
            elapsed = None
        if elapsed is not None and elapsed < DUPLICATE_SCAN_GAP_SECONDS:
            return {
                "ok": False,
                "message": f"Duplicate scan ignored. Try again after {int(DUPLICATE_SCAN_GAP_SECONDS - elapsed) + 1} seconds.",
                "student": student,
                "visit": None,
                "action": "duplicate",
            }

    open_visit = fetch_open_visit(student_id, today)

    if open_visit is None:
        visit = create_visit(student)
        save_visits(load_visits())
        return {
            "ok": True,
            "message": f"Entry saved: {student['name']} ({student['student_id']})",
            "student": student,
            "visit": visit,
            "action": "entry",
        }

    open_visit = update_visit_exit(student_id, today)
    if open_visit is None:
        return {
            "ok": False,
            "message": "Could not update the existing visit. Please try again.",
            "student": student,
            "visit": None,
            "action": "error",
        }
    save_visits(load_visits())
    return {
        "ok": True,
        "message": f"Exit saved: {student['name']} ({student['student_id']})",
        "student": student,
        "visit": open_visit,
        "action": "exit",
    }


def process_scan(student_id):
    result = process_scan_result(student_id)
    return result["ok"], result["message"]


def get_recent_visits(limit=10):
    visits = load_visits()
    recent = list(reversed(visits))
    if limit is not None:
        return recent[:limit]
    return recent


def get_active_visits():
    today = current_date_text()
    return [visit for visit in load_visits() if visit["date"] == today and not visit["exit_time"].strip()]


def get_dashboard_summary():
    students = load_students()
    visits = load_visits()
    today = current_date_text()
    active_visits = [visit for visit in visits if visit["date"] == today and not visit["exit_time"].strip()]
    today_visits = [visit for visit in visits if visit["date"] == today]

    return {
        "student_count": len(students),
        "total_visits": len(visits),
        "today_visits": len(today_visits),
        "inside_count": len(active_visits),
        "today": today,
    }


def with_student_details(visits, students=None):
    student_map = students if students is not None else load_students()
    return [
        {
            **visit,
            "course": student_map.get(visit["student_id"], {}).get("course", visit.get("course", "")),
            "father_name": student_map.get(visit["student_id"], {}).get("father_name", visit.get("father_name", "")),
        }
        for visit in visits
    ]


def build_daily_summary(visits):
    summary_map = {}
    for visit in visits:
        row = summary_map.setdefault(
            visit["date"],
            {"date": visit["date"], "total_visits": 0, "completed_visits": 0, "inside_count": 0},
        )
        row["total_visits"] += 1
        if visit["exit_time"].strip():
            row["completed_visits"] += 1
        else:
            row["inside_count"] += 1
    return sorted(summary_map.values(), key=lambda item: item["date"], reverse=True)


def build_weekly_summary(visits, today=None):
    summary_map = {}
    current_day = today or today_local()
    for visit in visits:
        visit_date = datetime.strptime(visit["date"], "%Y-%m-%d").date()
        iso_year, iso_week, iso_day = visit_date.isocalendar()
        week_key = f"{iso_year}-W{iso_week:02d}"
        start_date = visit_date - timedelta(days=iso_day - 1)
        end_date = start_date + timedelta(days=6)
        row = summary_map.setdefault(
            week_key,
            {
                "week_label": week_key,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "total_visits": 0,
                "completed_visits": 0,
                "inside_count": 0,
                "is_completed": end_date < current_day,
            },
        )
        row["total_visits"] += 1
        if visit["exit_time"].strip():
            row["completed_visits"] += 1
        else:
            row["inside_count"] += 1
    return [
        item
        for item in sorted(summary_map.values(), key=lambda item: item["week_label"], reverse=True)
        if item["is_completed"]
    ]


def build_dashboard_payload(recent_limit=12):
    students = load_students()
    visits = load_visits()
    today = current_date_text()
    recent_visits = [visit for visit in reversed(visits) if visit["date"] == today][:recent_limit]
    active_visits = [visit for visit in visits if visit["date"] == today and not visit["exit_time"].strip()]
    daily_summary = build_daily_summary(visits)
    weekly_summary = build_weekly_summary(visits)

    return {
        "summary": {
            "student_count": len(students),
            "total_visits": len(visits),
            "today_visits": sum(1 for visit in visits if visit["date"] == today),
            "inside_count": len(active_visits),
            "today": today,
        },
        "recent_visits": recent_visits,
        "recent_visits_with_students": with_student_details(recent_visits, students),
        "active_visits": with_student_details(active_visits, students),
        "daily_summary": daily_summary,
        "weekly_summary": weekly_summary,
    }


def _safe_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_date_filter(value):
    value = str(value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _week_label_from_date(date_text):
    try:
        visit_date = datetime.strptime(date_text, "%Y-%m-%d").date()
    except ValueError:
        return ""
    iso_year, iso_week, _ = visit_date.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"


def _mask_identifier(value):
    value = str(value or "").strip()
    if not value:
        return ""
    if len(value) <= 4:
        return "*" * len(value)
    return f"{value[:2]}{'*' * max(len(value) - 4, 1)}{value[-2:]}"


def _mask_person_name(value):
    value = " ".join(str(value or "").strip().split())
    if not value:
        return ""

    masked_parts = []
    for part in value.split(" "):
        if len(part) <= 2:
            masked_parts.append(part[0] + "*" * max(len(part) - 1, 0))
            continue
        masked_parts.append(f"{part[0]}{'*' * (len(part) - 2)}{part[-1]}")
    return " ".join(masked_parts)


def _format_duration(minutes):
    if minutes is None:
        return "-"
    hours, remainder = divmod(max(int(minutes), 0), 60)
    return f"{hours}h {remainder:02d}m" if hours else f"{remainder}m"


def _prepare_analytics_visit(visit, students, now=None):
    now = now or now_local()
    student = students.get(visit["student_id"], {})
    entry_timestamp = parse_timestamp(visit["date"], visit["entry_time"])
    exit_timestamp = parse_timestamp(visit["date"], visit["exit_time"])
    duration_minutes = None
    if entry_timestamp is not None:
        end_timestamp = exit_timestamp or now
        duration_minutes = max(int((end_timestamp - entry_timestamp).total_seconds() // 60), 0)

    return {
        **visit,
        "name": student.get("name", visit.get("name", "")),
        "father_name": student.get("father_name", visit.get("father_name", "")),
        "course": student.get("course", visit.get("course", "")),
        "phone": student.get("phone", ""),
        "status": "inside" if not visit["exit_time"].strip() else "completed",
        "week_label": _week_label_from_date(visit["date"]),
        "hour_bucket": f"{visit['entry_time'][:2]}:00" if visit["entry_time"] else "--:--",
        "duration_minutes": duration_minutes,
        "duration_label": _format_duration(duration_minutes),
    }


def _get_analytics_visits():
    global _ANALYTICS_CACHE

    signature = _analytics_source_signature()
    if signature and signature == _ANALYTICS_CACHE["signature"]:
        return _ANALYTICS_CACHE["visits"], _ANALYTICS_CACHE["token"]

    with _ANALYTICS_CACHE_LOCK:
        signature = _analytics_source_signature()
        if signature and signature == _ANALYTICS_CACHE["signature"]:
            return _ANALYTICS_CACHE["visits"], _ANALYTICS_CACHE["token"]

        students = load_students()
        visits = load_visits()
        now = now_local()
        prepared = [_prepare_analytics_visit(visit, students, now=now) for visit in visits]
        token = hashlib.sha1(signature.encode("utf-8")).hexdigest()[:12] if signature else "analytics-empty"
        _ANALYTICS_CACHE = {"signature": signature, "token": token, "visits": prepared}
        return prepared, token


def _apply_analytics_filters(visits, filters):
    date_from = _parse_date_filter(filters.get("date_from"))
    date_to = _parse_date_filter(filters.get("date_to"))
    week = str(filters.get("week", "")).strip()
    course = str(filters.get("course", "")).strip().casefold()
    state = str(filters.get("state", "all")).strip().casefold() or "all"
    search = str(filters.get("search", "")).strip().casefold()

    filtered = []
    for visit in visits:
        visit_date = _parse_date_filter(visit["date"])
        if date_from and (visit_date is None or visit_date < date_from):
            continue
        if date_to and (visit_date is None or visit_date > date_to):
            continue
        if week and visit["week_label"] != week:
            continue
        if course and visit.get("course", "").casefold() != course:
            continue
        if state == "inside" and visit["status"] != "inside":
            continue
        if state == "completed" and visit["status"] != "completed":
            continue
        if search:
            haystack = " ".join(
                [
                    visit.get("visit_id", ""),
                    visit.get("student_id", ""),
                    visit.get("name", ""),
                    visit.get("father_name", ""),
                    visit.get("course", ""),
                    visit.get("date", ""),
                    visit.get("entry_time", ""),
                    visit.get("exit_time", ""),
                ]
            ).casefold()
            if search not in haystack:
                continue
        filtered.append(visit)

    return filtered


def _hourly_distribution(visits):
    counter = Counter()
    for visit in visits:
        bucket = visit.get("hour_bucket") or "--:--"
        counter[bucket] += 1

    return [{"label": f"{hour:02d}:00", "value": counter.get(f"{hour:02d}:00", 0)} for hour in range(24)]


def _branch_distribution(visits, limit=8):
    counter = Counter()
    for visit in visits:
        branch = visit.get("course", "").strip() or "Unassigned"
        counter[branch] += 1
    return [{"label": label, "value": value} for label, value in counter.most_common(limit)]


def _series_from_counter(counter, sort_key=None):
    items = counter.items()
    if sort_key is None:
        items = sorted(items)
    else:
        items = sorted(items, key=sort_key)
    return [{"label": label, "value": value} for label, value in items]


def _daily_distribution(visits):
    counter = Counter()
    for visit in visits:
        counter[visit["date"]] += 1
    return _series_from_counter(counter)


def _weekly_distribution(visits):
    counter = Counter()
    for visit in visits:
        counter[visit["week_label"]] += 1
    return _series_from_counter(counter)


def _build_insights(visits, daily_series, hourly_series):
    inside_count = sum(1 for visit in visits if visit["status"] == "inside")
    completed_durations = [visit["duration_minutes"] for visit in visits if visit["status"] == "completed" and visit["duration_minutes"] is not None]

    busiest_day = max(daily_series, key=lambda item: item["value"], default=None)
    quiet_day = min([item for item in daily_series if item["value"] > 0], key=lambda item: item["value"], default=None)
    busiest_hour = max(hourly_series, key=lambda item: item["value"], default=None)
    quiet_hour = min([item for item in hourly_series if item["value"] > 0], key=lambda item: item["value"], default=None)

    return {
        "avg_duration_label": _format_duration(sum(completed_durations) // len(completed_durations)) if completed_durations else "-",
        "inside_count": inside_count,
        "busiest_day": busiest_day,
        "quiet_day": quiet_day,
        "busiest_hour": busiest_hour,
        "quiet_hour": quiet_hour,
        "irregular_open_visits": sum(1 for visit in visits if visit["status"] == "inside" and visit["date"] != current_date_text()),
    }


def _apply_visibility(visits, full_view=False):
    rows = []
    for visit in visits:
        row = dict(visit)
        if not full_view:
            row["student_id"] = _mask_identifier(row.get("student_id", ""))
            row["name"] = _mask_person_name(row.get("name", ""))
            row["father_name"] = _mask_person_name(row.get("father_name", ""))
            row["phone"] = _mask_identifier(row.get("phone", ""))
        rows.append(row)
    return rows


def build_analytics_payload(filters=None, view_mode="masked", row_limit=250, source_token=None, authenticated=False):
    filters = filters or {}
    visits, cache_token = _get_analytics_visits()
    if source_token and source_token == cache_token:
        return {
            "ok": True,
            "unchanged": True,
            "cache_token": cache_token,
            "view": {
                "requested": view_mode,
                "applied": "full" if authenticated and view_mode == "full" else "masked",
                "authenticated": authenticated,
                "can_view_full": authenticated,
            },
        }

    filtered = _apply_analytics_filters(visits, filters)
    filtered_desc = list(reversed(filtered))
    row_limit = max(25, min(_safe_int(row_limit, 250), 1000))
    applied_view = "full" if authenticated and view_mode == "full" else "masked"

    daily_series = _daily_distribution(filtered)
    weekly_series = _weekly_distribution(filtered)
    hourly_series = _hourly_distribution(filtered)
    branch_series = _branch_distribution(filtered)
    insights = _build_insights(filtered, daily_series, hourly_series)

    visible_rows = _apply_visibility(filtered_desc[:row_limit], full_view=applied_view == "full")
    courses = sorted({visit.get("course", "").strip() for visit in visits if visit.get("course", "").strip()})
    weeks = sorted({visit.get("week_label", "") for visit in visits if visit.get("week_label", "")}, reverse=True)

    return {
        "ok": True,
        "unchanged": False,
        "cache_token": cache_token,
        "generated_at": now_local().isoformat(),
        "view": {
            "requested": view_mode,
            "applied": applied_view,
            "authenticated": authenticated,
            "can_view_full": authenticated,
        },
        "filters": {
            "search": str(filters.get("search", "")).strip(),
            "course": str(filters.get("course", "")).strip(),
            "week": str(filters.get("week", "")).strip(),
            "date_from": str(filters.get("date_from", "")).strip(),
            "date_to": str(filters.get("date_to", "")).strip(),
            "state": str(filters.get("state", "all")).strip() or "all",
            "row_limit": row_limit,
            "course_options": courses,
            "week_options": weeks,
        },
        "summary": {
            "total_visits": len(filtered),
            "unique_students": len({visit["student_id"] for visit in filtered}),
            "inside_count": insights["inside_count"],
            "completed_count": sum(1 for visit in filtered if visit["status"] == "completed"),
            "avg_duration_label": insights["avg_duration_label"],
            "branch_count": len({visit.get("course", "").strip() for visit in filtered if visit.get("course", "").strip()}),
        },
        "insights": {
            "busiest_day": insights["busiest_day"],
            "quiet_day": insights["quiet_day"],
            "busiest_hour": insights["busiest_hour"],
            "quiet_hour": insights["quiet_hour"],
            "irregular_open_visits": insights["irregular_open_visits"],
        },
        "charts": {
            "daily": daily_series,
            "weekly": weekly_series,
            "hourly": hourly_series,
            "branches": branch_series,
        },
        "table": {
            "total_rows": len(filtered),
            "returned_rows": len(visible_rows),
            "rows": visible_rows,
        },
    }


def build_analytics_export_rows(filters=None, view_mode="masked", authenticated=False):
    filters = filters or {}
    visits, _ = _get_analytics_visits()
    filtered = list(reversed(_apply_analytics_filters(visits, filters)))
    applied_view = "full" if authenticated and view_mode == "full" else "masked"
    rows = _apply_visibility(filtered, full_view=applied_view == "full")
    return applied_view, rows
